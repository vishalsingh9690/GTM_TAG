from bs4 import BeautifulSoup 
from openpyxl import Workbook
import xlrd
import csv
import os


import requests
scr2 = "google_tag.script"
match_url = "https://www.googletagmanager.com/gtm.js"
print(os.path.dirname(os.path.abspath(__file__)))

# rb = xlrd.open_workbook("Galderma GTM.xlsx")
# sheet = rb.sheet_by_index(0)

# fieldnames = ['url', 'gtm']
# new_url_csv = open("tags2.csv", "w")
# out_writer = csv.DictWriter(new_url_csv, fieldnames=fieldnames)
# out_writer.writeheader()

# for row in range(2,sheet.nrows):
# 	print(sheet.cell(row,1).value)
# 	print(sheet.cell(row,0).value)
# 	out_writer.writerow({
# 		'url' : sheet.cell(row,0).value,
# 		'gtm' : sheet.cell(row,1).value
# 	})

fieldnames = ['url', 'gtm', 'status', 'error']
new_url_csv = open("output.csv", "w")
out_writer = csv.DictWriter(new_url_csv, fieldnames=fieldnames)
out_writer.writeheader()
mapped_url_csv = open("tags.csv", "r")
mapped_url_data = csv.DictReader(mapped_url_csv)
for csv_row in mapped_url_data:
	url = csv_row["url"]
	code = csv_row["gtm"]
	out_data = {
		"url" : url,
		"gtm" : code,
		"status" : "Unmatched",
		'error' : ''
	}
	try:
		response = requests.get(url)
		soup = BeautifulSoup(response.content)
		scripts = soup.find_all('script')
	except Exception as e:
		out_data.update({
			"status" : "Error",
			"error" : e.message
		})
		scripts = []
	for script in scripts:
		find_tag = script.get('src')
		if find_tag:
			if (find_tag.find(scr2) != -1):
				try:
					new_response = requests.get(url+find_tag)
					soup2 = BeautifulSoup(new_response.content)
					p_tag = soup2.find('p')
					p_data = p_tag.string
					if (p_data.find(match_url) != -1) and (p_data.find(code) != -1):
						out_data["status"] = "Matched"
				except Exception as e:
					out_data.update({
						"status" : "Error",
						"error" : e.message
					})
	out_writer.writerow(out_data)

mapped_url_csv.close()
new_url_csv.close()

new_url_csv = open("output.csv", "r")
new_url_csv_data = csv.DictReader(new_url_csv)
for csv_row in new_url_csv_data:
	if csv_row["status"] in ["Error", "Unmatched"]:
		assert False
		break



